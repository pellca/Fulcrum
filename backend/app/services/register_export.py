"""Register export: actions/commitments (+ chases/links) as an xlsx workbook or a csv zip.

Every query here is either a single bulk query or a joinedload — no per-row queries,
with one known exception: Link rows pointing at a "meeting" resolve their title via
the timeline TITLE_RESOLVERS getter, which dereferences r.forum lazily, so a batch of
meeting-typed link refs still costs one extra query per row (forum names aren't a
register export concern worth a bespoke eager-load option here).
Chase aggregates (count / last / next) come from one pass over the Chase table, and
Link target titles are otherwise resolved via one batched query per distinct entity
type (reusing the timeline TITLE_RESOLVERS model/getter mapping), not one per row.
"""

import csv
import io
import zipfile
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from ..models import Action, Chase, Commitment, Link
from .timeline import TITLE_RESOLVERS

# openpyxl raises IllegalCharacterError if a cell string contains a C0 control
# char (e.g. "\x07") -- strip those before they ever reach a cell.
def _clean(value):
    return ILLEGAL_CHARACTERS_RE.sub("", value) if isinstance(value, str) else value


# strings starting with one of these become live formulas when the file is
# opened in Excel (xlsx: openpyxl would otherwise infer data_type 'f'; csv:
# Excel evaluates on open) -- neutralize formula injection from user-entered
# titles/notes/etc.
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

ACTION_COLUMNS = [
    "id", "title", "description", "status", "priority", "owner", "owner_email",
    "commitment_id", "commitment_title", "workstream", "due_date", "created_at",
    "updated_at", "chase_count", "last_chased_on", "next_chase_on",
]
COMMITMENT_COLUMNS = [
    "id", "title", "description", "status", "priority", "origin", "origin_detail",
    "owner", "owner_email", "workstream", "due_date", "created_at", "updated_at",
    "action_count", "chase_count", "last_chased_on", "next_chase_on",
]
CHASE_COLUMNS = [
    "id", "target_type", "target_id", "target_title", "chased_on", "method", "note",
    "next_chase_on",
]
LINK_COLUMNS = [
    "id", "from_type", "from_id", "from_title", "to_type", "to_id", "to_title",
    "kind", "rationale",
]


def _iso(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _chase_stats(db: Session) -> dict[tuple[str, int], dict]:
    """(type, id) -> {count, last_chased_on, next_chase_on}. One query over every
    Chase row, ascending, so the final write per key lands on the most recent chase —
    same technique as services.chase.latest_chase_map."""
    stats: dict[tuple[str, int], dict] = {}
    for chase in db.query(Chase).order_by(Chase.chased_on, Chase.id):
        if chase.action_id is not None:
            key = ("action", chase.action_id)
        elif chase.commitment_id is not None:
            key = ("commitment", chase.commitment_id)
        else:
            continue
        entry = stats.setdefault(key, {"count": 0, "last_chased_on": None, "next_chase_on": None})
        entry["count"] += 1
        entry["last_chased_on"] = chase.chased_on
        entry["next_chase_on"] = chase.next_chase_on
    return stats


def _batch_resolve_titles(db: Session, refs: set[tuple[str, int]]) -> dict[tuple[str, int], str]:
    """Resolve every (type, id) ref using TITLE_RESOLVERS, at most one query per
    distinct entity type present in `refs` (not one per row)."""
    by_type: dict[str, set[int]] = {}
    for entity_type, entity_id in refs:
        by_type.setdefault(entity_type, set()).add(entity_id)

    titles: dict[tuple[str, int], str] = {}
    for entity_type, ids in by_type.items():
        entry = TITLE_RESOLVERS.get(entity_type)
        if not entry:
            for entity_id in ids:
                titles[(entity_type, entity_id)] = f"{entity_type} #{entity_id}"
            continue
        model, getter = entry
        found: set[int] = set()
        for row in db.query(model).filter(model.id.in_(ids)):
            titles[(entity_type, row.id)] = getter(row)
            found.add(row.id)
        for missing_id in ids - found:
            # only reachable for a Link whose target row is gone: actions/commitments
            # scrub their own links via services.bulk on delete, so this fallback is
            # really just for purged mail (mail purge doesn't touch Link rows)
            titles[(entity_type, missing_id)] = f"{entity_type} #{missing_id} (deleted)"
    return titles


def _action_rows(db: Session, chase_stats: dict) -> tuple[list[dict], dict[int, str]]:
    actions = (
        db.query(Action)
        .options(
            joinedload(Action.owner),
            joinedload(Action.workstream),
            joinedload(Action.commitment),
        )
        .order_by(Action.id)
        .all()
    )
    rows = []
    titles: dict[int, str] = {}
    for a in actions:
        titles[a.id] = a.title
        stats = chase_stats.get(("action", a.id), {})
        rows.append(
            {
                "id": a.id,
                "title": a.title,
                "description": a.description,
                "status": a.status,
                "priority": a.priority,
                "owner": a.owner.name if a.owner else None,
                "owner_email": a.owner.email if a.owner else None,
                "commitment_id": a.commitment_id,
                "commitment_title": a.commitment.title if a.commitment else None,
                "workstream": a.workstream.name if a.workstream else None,
                "due_date": _iso(a.due_date),
                "created_at": _iso(a.created_at),
                "updated_at": _iso(a.updated_at),
                "chase_count": stats.get("count", 0),
                "last_chased_on": _iso(stats.get("last_chased_on")),
                "next_chase_on": _iso(stats.get("next_chase_on")),
            }
        )
    return rows, titles


def _commitment_rows(db: Session, chase_stats: dict) -> tuple[list[dict], dict[int, str]]:
    commitments = (
        db.query(Commitment)
        .options(
            joinedload(Commitment.owner),
            joinedload(Commitment.workstream),
            joinedload(Commitment.actions),
        )
        .order_by(Commitment.id)
        .all()
    )
    rows = []
    titles: dict[int, str] = {}
    for c in commitments:
        titles[c.id] = c.title
        stats = chase_stats.get(("commitment", c.id), {})
        rows.append(
            {
                "id": c.id,
                "title": c.title,
                "description": c.description,
                "status": c.status,
                "priority": c.priority,
                "origin": c.origin,
                "origin_detail": c.origin_detail,
                "owner": c.owner.name if c.owner else None,
                "owner_email": c.owner.email if c.owner else None,
                "workstream": c.workstream.name if c.workstream else None,
                "due_date": _iso(c.due_date),
                "created_at": _iso(c.created_at),
                "updated_at": _iso(c.updated_at),
                "action_count": len(c.actions),
                "chase_count": stats.get("count", 0),
                "last_chased_on": _iso(stats.get("last_chased_on")),
                "next_chase_on": _iso(stats.get("next_chase_on")),
            }
        )
    return rows, titles


def _chase_rows(db: Session, action_titles: dict[int, str], commitment_titles: dict[int, str]) -> list[dict]:
    chases = db.query(Chase).order_by(Chase.chased_on.desc(), Chase.id.desc()).all()
    rows = []
    for chase in chases:
        if chase.action_id is not None:
            target_type, target_id = "action", chase.action_id
            target_title = action_titles.get(target_id)
        elif chase.commitment_id is not None:
            target_type, target_id = "commitment", chase.commitment_id
            target_title = commitment_titles.get(target_id)
        else:
            continue
        rows.append(
            {
                "id": chase.id,
                "target_type": target_type,
                "target_id": target_id,
                "target_title": target_title,
                "chased_on": _iso(chase.chased_on),
                "method": chase.method,
                "note": chase.note,
                "next_chase_on": _iso(chase.next_chase_on),
            }
        )
    return rows


def _link_rows(db: Session) -> list[dict]:
    links = (
        db.query(Link)
        .filter(
            Link.from_type.in_(("action", "commitment")) | Link.to_type.in_(("action", "commitment"))
        )
        .order_by(Link.id)
        .all()
    )
    refs: set[tuple[str, int]] = set()
    for link in links:
        refs.add((link.from_type, link.from_id))
        refs.add((link.to_type, link.to_id))
    titles = _batch_resolve_titles(db, refs)

    rows = []
    for link in links:
        rows.append(
            {
                "id": link.id,
                "from_type": link.from_type,
                "from_id": link.from_id,
                "from_title": titles.get((link.from_type, link.from_id)),
                "to_type": link.to_type,
                "to_id": link.to_id,
                "to_title": titles.get((link.to_type, link.to_id)),
                "kind": link.kind,
                "rationale": link.rationale,
            }
        )
    return rows


def _write_xlsx(sheets: list[tuple[str, list[str], list[dict]]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    for name, columns, rows in sheets:
        ws = wb.create_sheet(name)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = bold
        ws.freeze_panes = "A2"
        for row in rows:
            ws.append([_clean(row.get(col)) for col in columns])
            for cell in ws[ws.max_row]:
                if isinstance(cell.value, str):
                    # force the cell to a plain string type even if the value
                    # starts with =/+/-/@ etc, so openpyxl never writes it as
                    # a live formula (data_type 'f')
                    cell.data_type = "s"
        for idx, col in enumerate(columns, start=1):
            lengths = [len(col)] + [
                len(str(row.get(col))) for row in rows if row.get(col) is not None
            ]
            ws.column_dimensions[get_column_letter(idx)].width = min(max(lengths) + 2, 40)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _csv_value(value):
    value = _clean(value)
    if isinstance(value, str) and value[:1] in FORMULA_PREFIXES:
        # leading single-quote is CSV/Excel's own escape for "treat as text,
        # not a formula" -- it's stripped back off on display
        return "'" + value
    return value


def _write_csv_zip(sheets: list[tuple[str, list[str], list[dict]]]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, columns, rows in sheets:
            text_buffer = io.StringIO()
            writer = csv.writer(text_buffer)
            writer.writerow(columns)
            for row in rows:
                writer.writerow([_csv_value(row.get(col)) for col in columns])
            zf.writestr(f"{name.lower()}.csv", text_buffer.getvalue().encode("utf-8-sig"))
    return buffer.getvalue()


def build_export(db: Session, format: str, chases: bool, links: bool) -> tuple[bytes, str, str]:
    """Returns (content_bytes, media_type, file_extension)."""
    chase_stats = _chase_stats(db)
    action_rows, action_titles = _action_rows(db, chase_stats)
    commitment_rows, commitment_titles = _commitment_rows(db, chase_stats)

    sheets = [
        ("Actions", ACTION_COLUMNS, action_rows),
        ("Commitments", COMMITMENT_COLUMNS, commitment_rows),
    ]
    if chases:
        sheets.append(("Chases", CHASE_COLUMNS, _chase_rows(db, action_titles, commitment_titles)))
    if links:
        sheets.append(("Links", LINK_COLUMNS, _link_rows(db)))

    if format == "xlsx":
        return (
            _write_xlsx(sheets),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    return _write_csv_zip(sheets), "application/zip", "zip"
