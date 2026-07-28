"""Import actions/commitments from CSV (Copilot extraction, generic template) or
an MS Planner Excel export.

Planner's "Export plan to Excel" sheet has columns like:
Task Name, Bucket Name, Progress, Priority, Assigned To, Due Date, Description, ...
Generic/Copilot CSVs use lowercase snake_case headers and may carry a `type` column
(action | commitment) plus `origin` for commitments.
"""

import csv
import io
import re
from datetime import date, datetime

from sqlalchemy.orm import Session

from ..models import Action, Commitment, Forum, Link, Meeting
from .quickadd import ORIGINS, find_person, find_workstream

PROGRESS_MAP = {
    "not started": "todo",
    "in progress": "in_progress",
    "completed": "done",
}
PRIORITY_MAP = {
    "urgent": "high",
    "important": "high",
    "high": "high",
    "medium": "medium",
    "low": "low",
}

# canonical field -> accepted header spellings (lowercased)
HEADER_ALIASES = {
    "type": ["type", "item type"],
    "title": ["task name", "title", "name"],
    "description": ["description", "notes", "context"],
    "owner": ["assigned to", "owner", "assignee"],
    "due": ["due date", "due", "due_date"],
    "status": ["progress", "status"],
    "priority": ["priority"],
    "workstream": ["bucket name", "workstream", "bucket"],
    "origin": ["origin"],
    "meeting": ["meeting", "source meeting", "forum"],
}


def _parse_date(value: str) -> date | None:
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d %B %Y", "%d-%b-%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value).date()
    except ValueError:
        return None


def _map_headers(headers: list[str]) -> dict[str, int]:
    """canonical field -> column index."""
    mapping: dict[str, int] = {}
    lowered = [h.strip().lower() for h in headers]
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in lowered:
                mapping[field] = lowered.index(alias)
                break
    return mapping


def rows_from_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    return [row for row in csv.reader(io.StringIO(text)) if any(cell.strip() for cell in row)]


def rows_from_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if cell is None else str(cell) for cell in row]
        if any(cell.strip() for cell in cells):
            rows.append(cells)
    workbook.close()
    return rows


def find_meeting(db: Session, value: str) -> Meeting | None:
    """Resolve 'AET Weekly 2026-08-03' (or just a forum name) to a meeting.

    With a date: that forum's meeting closest to the date. Without: closest to today.
    """
    value = (value or "").strip()
    if not value:
        return None
    target = None
    date_match = re.search(r"(\d{4}-\d{2}-\d{2})|(\d{1,2}/\d{1,2}/\d{4})", value)
    if date_match:
        target = _parse_date(date_match.group(0))
        value = value.replace(date_match.group(0), "").strip(" -–—,")
    forum = db.query(Forum).filter(Forum.name.ilike(f"%{value}%")).first() if value else None
    if forum is None:
        return None
    meetings = db.query(Meeting).filter(Meeting.forum_id == forum.id).all()
    if not meetings:
        return None
    anchor = target or date.today()
    return min(meetings, key=lambda m: abs((m.scheduled_at.date() - anchor).days))


def _meeting_label(meeting: Meeting) -> str:
    return f"{meeting.forum.name} — {meeting.scheduled_at:%d %b %Y}"


def preview_import(db: Session, rows: list[list[str]]) -> dict:
    """Parse rows into proposed actions/commitments without writing anything."""
    if not rows:
        return {"columns": {}, "items": [], "skipped": 0}
    mapping = _map_headers(rows[0])
    if "title" not in mapping:
        raise ValueError(
            "Could not find a title column (expected one of: "
            + ", ".join(HEADER_ALIASES["title"])
            + ")"
        )
    items, skipped = [], 0
    for row in rows[1:]:
        def cell(field: str) -> str:
            index = mapping.get(field)
            return row[index].strip() if index is not None and index < len(row) else ""

        title = cell("title")
        if not title:
            skipped += 1
            continue
        item_type = cell("type").lower()
        if item_type not in ("action", "commitment"):
            item_type = "action"
        owner_name = cell("owner").split(";")[0].strip()
        owner = find_person(db, owner_name) if owner_name else None
        ws_name = cell("workstream")
        workstream = find_workstream(db, ws_name) if ws_name else None
        origin = cell("origin").lower()
        meeting_ref = cell("meeting")
        meeting = find_meeting(db, meeting_ref) if meeting_ref else None
        items.append(
            {
                "type": item_type,
                "meeting_id": meeting.id if meeting else None,
                "meeting_label": _meeting_label(meeting) if meeting else (meeting_ref or None),
                "meeting_matched": meeting is not None or not meeting_ref,
                "title": title,
                "description": cell("description") or None,
                "owner_id": owner.id if owner else None,
                "owner_name": owner.name if owner else (owner_name or None),
                "owner_matched": owner is not None or not owner_name,
                "workstream_id": workstream.id if workstream else None,
                "workstream_name": workstream.name if workstream else (ws_name or None),
                "due_date": (d := _parse_date(cell("due"))) and d.isoformat(),
                "status": PROGRESS_MAP.get(cell("status").lower(), "todo"),
                "priority": PRIORITY_MAP.get(cell("priority").lower(), "medium"),
                "origin": origin if origin in ORIGINS else "principal",
            }
        )
    return {"columns": {k: rows[0][v] for k, v in mapping.items()}, "items": items, "skipped": skipped}


def commit_import(db: Session, items: list[dict], default_meeting_id: int | None = None) -> dict:
    created = {"actions": 0, "commitments": 0, "meeting_links": 0}
    for item in items:
        common = dict(
            title=item["title"],
            description=item.get("description"),
            owner_id=item.get("owner_id"),
            workstream_id=item.get("workstream_id"),
            due_date=date.fromisoformat(item["due_date"]) if item.get("due_date") else None,
            priority=item.get("priority", "medium"),
        )
        if item.get("type") == "commitment":
            row = Commitment(**common, origin=item.get("origin", "principal"))
            created["commitments"] += 1
        else:
            row = Action(**common, status=item.get("status", "todo"))
            created["actions"] += 1
        db.add(row)

        meeting_id = item.get("meeting_id") or default_meeting_id
        if meeting_id and db.get(Meeting, meeting_id):
            db.flush()
            db.add(
                Link(
                    from_type="meeting",
                    from_id=meeting_id,
                    to_type=item.get("type", "action"),
                    to_id=row.id,
                    kind="informs",
                    rationale="Raised in this meeting (imported)",
                )
            )
            created["meeting_links"] += 1
    db.commit()
    return created


TEMPLATES = {
    "actions": ["type", "title", "description", "owner", "workstream", "due", "status", "priority", "meeting"],
    "commitments": ["type", "title", "description", "owner", "workstream", "due", "origin", "priority", "meeting"],
    "topics": ["title", "description", "intent", "duration_minutes", "owner", "workstream", "due"],
    "key_dates": ["title", "due", "kind", "hard", "workstream", "description"],
}


def template_csv(name: str) -> str:
    headers = TEMPLATES[name]
    return ",".join(headers) + "\n"
