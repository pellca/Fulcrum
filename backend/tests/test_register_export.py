import csv
import io
import zipfile
from datetime import date, timedelta

import openpyxl

TODAY = date.today()


def _person(client, name, email=None):
    body = {"name": name}
    if email is not None:
        body["email"] = email
    return client.post("/api/people", json=body).json()


def _workstream(client, name):
    return client.post("/api/workstreams", json={"name": name}).json()


def _action(client, title, **kwargs):
    body = {"title": title, **kwargs}
    return client.post("/api/actions", json=body).json()


def _commitment(client, title, **kwargs):
    body = {"title": title, **kwargs}
    return client.post("/api/commitments", json=body).json()


def _chase(client, action_id=None, commitment_id=None, chased_on=None, next_chase_on=None, note=None):
    body = {
        "chased_on": chased_on or TODAY.isoformat(),
        "next_chase_on": next_chase_on,
        "note": note,
    }
    if action_id is not None:
        body["action_id"] = action_id
    if commitment_id is not None:
        body["commitment_id"] = commitment_id
    resp = client.post("/api/chases", json=body)
    assert resp.status_code == 201, resp.text
    return resp.json()


def _link(client, from_type, from_id, to_type, to_id, kind="relates", rationale=None):
    resp = client.post(
        "/api/links",
        json={
            "from_type": from_type, "from_id": from_id,
            "to_type": to_type, "to_id": to_id,
            "kind": kind, "rationale": rationale,
        },
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _seed(client):
    owner = _person(client, "Priya Shah", "priya.shah@bank.com")
    workstream = _workstream(client, "Payments Remediation")
    commitment = _commitment(
        client, "Deliver remediation report", owner_id=owner["id"], workstream_id=workstream["id"],
        due_date=(TODAY + timedelta(days=10)).isoformat(),
    )
    action = _action(
        client, "Draft remediation report", owner_id=owner["id"], workstream_id=workstream["id"],
        commitment_id=commitment["id"], due_date=(TODAY + timedelta(days=3)).isoformat(),
    )
    chase = _chase(
        client, action_id=action["id"],
        chased_on=(TODAY - timedelta(days=2)).isoformat(),
        next_chase_on=(TODAY + timedelta(days=5)).isoformat(),
        note="Chased Priya by email",
    )
    link = _link(client, "action", action["id"], "commitment", commitment["id"], kind="relates",
                 rationale="Delivery action for the commitment")
    return {"owner": owner, "workstream": workstream, "commitment": commitment,
            "action": action, "chase": chase, "link": link}


# ---------- xlsx ----------


def test_export_xlsx_default_sheets_and_flags(client):
    _seed(client)

    resp = client.get("/api/register/export?format=xlsx")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"fulcrum-register-{TODAY.strftime('%Y%m%d')}.xlsx"
    assert f'filename="{filename}"' in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) == {"Actions", "Commitments"}

    resp_full = client.get("/api/register/export?format=xlsx&chases=true&links=true")
    wb_full = openpyxl.load_workbook(io.BytesIO(resp_full.content))
    assert set(wb_full.sheetnames) == {"Actions", "Commitments", "Chases", "Links"}


def test_export_xlsx_header_and_known_row(client):
    seed = _seed(client)
    resp = client.get("/api/register/export?format=xlsx&chases=true&links=true")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))

    actions_ws = wb["Actions"]
    header = [cell.value for cell in actions_ws[1]]
    assert header == [
        "id", "title", "description", "status", "priority", "owner", "owner_email",
        "commitment_id", "commitment_title", "workstream", "due_date", "created_at",
        "updated_at", "chase_count", "last_chased_on", "next_chase_on",
    ]
    assert all(cell.font.bold for cell in actions_ws[1])
    assert actions_ws.freeze_panes == "A2"

    row = {header[i]: cell.value for i, cell in enumerate(actions_ws[2])}
    action = seed["action"]
    assert row["id"] == action["id"]
    assert row["title"] == "Draft remediation report"
    assert row["status"] == "todo"
    assert row["owner"] == "Priya Shah"
    assert row["owner_email"] == "priya.shah@bank.com"
    assert row["commitment_id"] == seed["commitment"]["id"]
    assert row["commitment_title"] == "Deliver remediation report"
    assert row["workstream"] == "Payments Remediation"
    assert row["due_date"] == (TODAY + timedelta(days=3)).isoformat()
    assert row["chase_count"] == 1
    assert row["last_chased_on"] == (TODAY - timedelta(days=2)).isoformat()
    assert row["next_chase_on"] == (TODAY + timedelta(days=5)).isoformat()

    commitments_ws = wb["Commitments"]
    c_header = [cell.value for cell in commitments_ws[1]]
    c_row = {c_header[i]: cell.value for i, cell in enumerate(commitments_ws[2])}
    assert c_row["title"] == "Deliver remediation report"
    assert c_row["action_count"] == 1
    assert c_row["owner"] == "Priya Shah"
    assert c_row["workstream"] == "Payments Remediation"


def test_export_xlsx_chases_and_links_content(client):
    seed = _seed(client)
    resp = client.get("/api/register/export?format=xlsx&chases=true&links=true")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))

    chases_ws = wb["Chases"]
    c_header = [cell.value for cell in chases_ws[1]]
    assert c_header == [
        "id", "target_type", "target_id", "target_title", "chased_on", "method", "note",
        "next_chase_on",
    ]
    c_row = {c_header[i]: cell.value for i, cell in enumerate(chases_ws[2])}
    assert c_row["id"] == seed["chase"]["id"]
    assert c_row["target_type"] == "action"
    assert c_row["target_id"] == seed["action"]["id"]
    assert c_row["target_title"] == "Draft remediation report"
    assert c_row["note"] == "Chased Priya by email"

    links_ws = wb["Links"]
    l_header = [cell.value for cell in links_ws[1]]
    assert l_header == [
        "id", "from_type", "from_id", "from_title", "to_type", "to_id", "to_title",
        "kind", "rationale",
    ]
    l_row = {l_header[i]: cell.value for i, cell in enumerate(links_ws[2])}
    assert l_row["from_type"] == "action"
    assert l_row["from_id"] == seed["action"]["id"]
    assert l_row["from_title"] == "Draft remediation report"
    assert l_row["to_type"] == "commitment"
    assert l_row["to_id"] == seed["commitment"]["id"]
    assert l_row["to_title"] == "Deliver remediation report"
    assert l_row["kind"] == "relates"


# ---------- csv zip ----------


def test_export_csv_zip_members_and_flags(client):
    _seed(client)

    resp = client.get("/api/register/export?format=csv")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/zip"
    filename = f"fulcrum-register-{TODAY.strftime('%Y%m%d')}.zip"
    assert f'filename="{filename}"' in resp.headers["content-disposition"]

    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    assert set(zf.namelist()) == {"actions.csv", "commitments.csv"}

    resp_full = client.get("/api/register/export?format=csv&chases=true&links=true")
    zf_full = zipfile.ZipFile(io.BytesIO(resp_full.content))
    assert set(zf_full.namelist()) == {"actions.csv", "commitments.csv", "chases.csv", "links.csv"}


def test_export_csv_zip_bom_and_known_row(client):
    seed = _seed(client)
    resp = client.get("/api/register/export?format=csv&chases=true&links=true")
    zf = zipfile.ZipFile(io.BytesIO(resp.content))

    raw = zf.read("actions.csv")
    assert raw.startswith(b"\xef\xbb\xbf")  # utf-8-sig BOM

    text = raw.decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text)))
    assert len(rows) == 1
    row = rows[0]
    assert row["id"] == str(seed["action"]["id"])
    assert row["title"] == "Draft remediation report"
    assert row["owner"] == "Priya Shah"
    assert row["owner_email"] == "priya.shah@bank.com"
    assert row["commitment_title"] == "Deliver remediation report"
    assert row["chase_count"] == "1"
    assert row["next_chase_on"] == (TODAY + timedelta(days=5)).isoformat()

    chases_text = zf.read("chases.csv").decode("utf-8-sig")
    chase_rows = list(csv.DictReader(io.StringIO(chases_text)))
    assert chase_rows[0]["target_title"] == "Draft remediation report"
    assert chase_rows[0]["note"] == "Chased Priya by email"

    links_text = zf.read("links.csv").decode("utf-8-sig")
    link_rows = list(csv.DictReader(io.StringIO(links_text)))
    assert link_rows[0]["from_title"] == "Draft remediation report"
    assert link_rows[0]["to_title"] == "Deliver remediation report"


# ---------- misc ----------


def test_export_bad_format_422(client):
    resp = client.get("/api/register/export?format=pdf")
    assert resp.status_code == 422


def test_export_empty_register_still_returns_workbook(client):
    resp = client.get("/api/register/export?format=xlsx&chases=true&links=true")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb["Actions"].max_row == 1  # header only
    assert [cell.value for cell in wb["Actions"][1]][0] == "id"


# ---------- illegal characters / formula injection ----------


def test_export_xlsx_strips_illegal_control_chars(client):
    # a C0 control char like \x07 (BEL) used to make openpyxl raise
    # IllegalCharacterError and 500 the whole export
    _action(client, "Bell\x07 in title")
    resp = client.get("/api/register/export?format=xlsx")
    assert resp.status_code == 200, resp.text
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    titles = [row[1].value for row in wb["Actions"].iter_rows(min_row=2)]
    assert "Bell in title" in titles
    assert not any("\x07" in (t or "") for t in titles)


def test_export_csv_strips_illegal_control_chars(client):
    _action(client, "Bell\x07 in title")
    resp = client.get("/api/register/export?format=csv")
    assert resp.status_code == 200, resp.text
    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    text = zf.read("actions.csv").decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text)))
    assert rows[0]["title"] == "Bell in title"


def test_export_xlsx_neutralizes_formula_titles(client):
    _action(client, "=HYPERLINK(\"http://evil.example\",\"click\")")
    resp = client.get("/api/register/export?format=xlsx")
    assert resp.status_code == 200, resp.text
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    cell = wb["Actions"]["B2"]
    assert cell.value.startswith("=HYPERLINK")
    assert cell.data_type != "f"


def test_export_csv_neutralizes_formula_titles(client):
    _action(client, "=HYPERLINK(\"http://evil.example\",\"click\")")
    resp = client.get("/api/register/export?format=csv")
    assert resp.status_code == 200, resp.text
    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    text = zf.read("actions.csv").decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text)))
    assert rows[0]["title"].startswith("'=HYPERLINK")
